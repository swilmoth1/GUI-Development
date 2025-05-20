import tkinter as tk
import customtkinter as ctk
from tkinter import ttk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
import math
import os
import json
from Segmentation_Settings_GUI import SegmentationSettingsGUI
from Recording_settings_GUI import RecordingsettingsGUI
from Annotation_Settings_GUI import AnnotationSettingsGUI
from Material_Defaults_GUI import MaterialDefaultsGUI
from Graph_Settings_GUI import GraphSettingsGUI
from PIL import Image, ImageTk
from collections import deque
import time
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import socket
import threading
import queue
from pypylon import pylon
import sys
import cv2
import xml.etree.ElementTree as ET
from ultralytics import YOLO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
import torch
import torch.nn.functional as F
sys.path.append(os.path.dirname(__file__))  # Ensure current folder is in path
from annotate_cy import annotate_raw_image




# Set theme and color
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# Configuration
NUM_GRAPHS = 5  # Updated to 5 graphs
HISTORY_LENGTH = 100  # Number of frames to show in history
ASSETS_DIR = os.path.join(os.path.dirname(__file__), "assets")
DEFAULT_IMG = os.path.join(ASSETS_DIR, "placeholder.png")

global output_data
global frame_count
global line_refs
global fill_refs
fill_refs = {} 
line_refs = {}
centerline_refs = {}
frame_count = 0


global cam
cam = 0
cam_lock = threading.Lock()
R_IP = "192.168.1.25"
R_PORT = 59152

global start_time_integer
start_time_integer = 0

# Camera IP and port
IP = "169.254.235.230"
PORT = 139

# Ensure assets directory exists
if not os.path.exists(ASSETS_DIR):
    os.makedirs(ASSETS_DIR)
global image_paths
image_paths = {
    "video_raw": os.path.join(ASSETS_DIR, "Sample Raw Image.png"),
    "video_annotated": os.path.join(ASSETS_DIR, "Sample Annotated Image.png"),
    "video_segmented": os.path.join(ASSETS_DIR, "Sample Segmented Image.png"),
    "image_raw": os.path.join(ASSETS_DIR, "Sample Raw Image 4.png"),
    "image_annotated": os.path.join(ASSETS_DIR, "Sample Annotated Image.png"),
    "image_segmented": os.path.join(ASSETS_DIR, "Sample Segmented Image.png")
}

preview_frames = {}
preview_labels = {}
preview_texts = {}

def load_image_preview(path, size=(160,120)):
    try:
        # Try to load specified image
        img = Image.open(path)
    except Exception as e:
        print(f"Error loading image {path}: {e}")
        # Create a blank placeholder if image doesn't exist
        img = Image.new('RGB', size, color='gray')
        # Add text to indicate placeholder
        from PIL import ImageDraw
        draw = ImageDraw.Draw(img)
        draw.text((10, size[1]//2), "No Preview", fill='white')
        # Save placeholder for future use
        img.save(path)
    
    try:
        img = img.resize(size, Image.Resampling.LANCZOS)
    except Exception as e:
        print(f"Error resizing image: {e}")
        img = img.resize(size, Image.NEAREST)
        
    return ctk.CTkImage(light_image=img, dark_image=img, size=size)

def on_closing():
    """Handle cleanup when main window is closed"""
    # Close all matplotlib figures
    plt.close('all')
    
    # Destroy any child windows
    for widget in window.winfo_children():
        if isinstance(widget, ctk.CTkToplevel):
            widget.destroy()
    
    # Cancel any pending after callbacks
    window.after_cancel(window.after_id)
    
    # Quit and destroy main window
    window.quit()
    window.destroy()

# Main GUI Window   
window = ctk.CTk()
window.title("Deposition Monitoring System")
window.protocol("WM_DELETE_WINDOW", on_closing)

# Store the first after callback ID
window.after_id = window.after(10, lambda: None)

#Testing messages
alert_message = tk.StringVar()
alert_message.set('Test')

# Maximize window while keeping controls
window.state('zoomed')  # For Windows

# Configure the main window's grid
window.grid_columnconfigure(0, weight=1)
window.grid_rowconfigure(1, weight=3)  # Image frame gets 3 parts
window.grid_rowconfigure(2, weight=3)  # Graph frame gets 3 parts
window.grid_rowconfigure(0, weight=1)  # Control panel gets 1 part

# Control panel field
control_frame = ctk.CTkFrame(master=window)
control_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=5)
control_frame.grid_columnconfigure(0, weight=4)  # Main buttons section gets more space
#control_frame.grid_columnconfigure(1, weight=1)  # Status section gets less space
control_title = ctk.CTkLabel(control_frame, text="Control Panel", font=("Arial", 12, "bold"))
control_title.grid(row=0, column=0, columnspan=2, sticky="w", padx=10, pady=5)

###### LOAD SETTINGS FROM BUTTON SELECTIONS

#read the json file for the Graph Settings
if os.path.exists("graph_settings.json"):
    with open("graph_settings.json", 'r') as f:
        global graph_settings
        graph_settings = json.load(f)

if os.path.exists("annotation_settings.json"):
    with open("annotation_settings.json", 'r') as f:
        annotation_settings = json.load(f)

if os.path.exists("class_values.json"):
    with open("class_values.json", 'r') as f:
        global class_values
        class_values = json.load(f)

if os.path.exists("material_defaults.json"):
    with open("material_defaults.json", 'r') as f:
        material_defaults = json.load(f)

if os.path.exists("recording_settings.json"):
    with open("recording_settings.json", 'r') as f:
        global recording_settings
        recording_settings = json.load(f)

if os.path.exists("segmentation_settings.json"):
    with open("segmentation_settings.json", 'r') as f:
        global segmentation_settings
        segmentation_settings = json.load(f)

if os.path.exists("raw_experiment_excel_data.json"):
    with open("raw_experiment_excel_data.json", 'r') as f:
        global raw_experiment_excel_data_settings
        raw_experiment_excel_data_settings = json.load(f)
        

# Buttons in Control Panel (stays in column 0)
buttons_frame = ctk.CTkFrame(control_frame)
buttons_frame.grid(row=1, column=0, padx=10, pady=5, sticky="w")

###############
def open_segmentation_settings():
    SegmentationSettingsGUI(window, callback = update_class_values)

def open_recording_settings():
    RecordingsettingsGUI(window, callback=update_gui_from_settings)
    
def open_annotation_settings():
    AnnotationSettingsGUI(window)
    
def open_material_defaults_settings():
    MaterialDefaultsGUI(window)

def open_graph_settings():
    GraphSettingsGUI(window, callback = update_graphs)

def set_material_defaults(choice=None):
    if choice == "Select Material":
        return
    else:
        material_selection = choice

seg_settings_button = ctk.CTkButton(buttons_frame, 
                                  text="Segmentation Settings",
                                  command=open_segmentation_settings)
seg_settings_button.grid(row=0, column=0, padx=10, pady=5)

recording_settings_button = ctk.CTkButton(buttons_frame, text="Recording Settings", command=open_recording_settings)
recording_settings_button.grid(row=0, column=1, padx=10, pady=5)

Annotation_settings_button = ctk.CTkButton(buttons_frame, 
                                         text="Annotation Settings",
                                         command=open_annotation_settings)
Annotation_settings_button.grid(row=0, column=2, padx=10, pady=5)

Graph_settings_button = ctk.CTkButton(buttons_frame, text="Graph Settings", command=open_graph_settings)
Graph_settings_button.grid(row=0, column=3, padx=10, pady=5)

Material_defaults_button = ctk.CTkButton(buttons_frame, text="Material Defaults", command=open_material_defaults_settings)
Material_defaults_button.grid(row=0, column=4, padx=10, pady=5)

# Set default material selection
material_selection = list(material_defaults.keys())[0]

Material_selection_button = ctk.CTkComboBox(buttons_frame, values=["Select Material"]+list(material_defaults.keys()),command=set_material_defaults)
Material_selection_button.grid(row=0, column=5, padx=10, pady=5)
###############

is_recording = False

### TOGGLE RECORDING FUNCTION. THIS IS WHERE THE RECORDING OF THE IMAGES STARTS.
def toggle_recording():
    global is_recording, timestamp, experiment_folder

    if record_button.cget("text") == "Start Recording":
        print("\n--- Starting Recording ---")
        t0 = time.perf_counter()

        record_button.configure(
            text="Stop Recording", 
            fg_color="red",
            hover_color="dark red"
        )
        is_recording = True
        t1 = time.perf_counter()
        print(f"[Timing] Button updated: {t1 - t0:.4f} s")

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        experiment_folder = os.path.join(recording_settings["recording_save_location"], timestamp)
        os.makedirs(experiment_folder, exist_ok=True)
        t2 = time.perf_counter()
        print(f"[Timing] Experiment folder created: {t2 - t1:.4f} s")

        threading.Thread(target=video_acquiring_worker, daemon=True).start()
        t3 = time.perf_counter()
        print(f"[Timing] Video acquiring thread started: {t3 - t2:.4f} s")

        threading.Thread(target=image_saving_worker, daemon=True).start()
        t4 = time.perf_counter()
        print(f"[Timing] Image saving thread started: {t4 - t3:.4f} s")

        threading.Thread(target=measurement_saving_worker, daemon=True).start()
        t5 = time.perf_counter()
        print(f"[Timing] Measurement saving thread started: {t5 - t4:.4f} s")

        threading.Thread(target=draw_preview_worker, daemon=True).start()
        t6 = time.perf_counter()
        print(f"[Timing] Preview drawing thread started: {t6 - t5:.4f} s")

        update_recording_loop()
        t7 = time.perf_counter()
        print(f"[Timing] update_recording_loop() scheduled: {t7 - t6:.4f} s")
        print(f"[Total Startup Time] {t7 - t0:.4f} s")

    else:
        print("\n--- Stopping Recording ---")
        record_button.configure(
            text="Start Recording", 
            fg_color="green",
            hover_color="dark green"
        )
        is_recording = False

        save_all_charts(graph_settings, canvases)
        update_camera_status(big_status_display,'idle')

def save_measurements_to_excel(frame_counter, measurements, excel_settings, exposure_time):
    global annotation_settings
    base_dir = experiment_folder
    excel_path = os.path.join(base_dir, "Raw Segmentation Data.xlsx")  # Updated filename

    # Construct row data from settings and measurements
    row_data = {
        "Frame": frame_counter,
        "Segmentation Type": excel_settings.get("Segmentation Type", ""),
        "Filename": f"frame_{frame_counter:05d}.png",
        "Starting Exposure Time": excel_settings.get("Starting Exposure Time (\u03bcs)", ""),
        "Ending Exposure Time": excel_settings.get("Ending Exposure Time (\u03bcs)", ""),
        "Increment": excel_settings.get("Increment (\u03bcs)", ""),
        "Exposure Time": exposure_time,
        "Filter Applied": excel_settings.get("Filter Applied", ""),
        "Material": excel_settings.get("Material", ""),
        "Job Number": excel_settings.get("Job Number", ""),
        "Wire Feed Speed (in/min)": excel_settings.get("Wire Feed Speed (in/min)", ""),
        "Travel Speed (in/min)": excel_settings.get("Travel Speed (in/min)", ""),
        "Camera Model": excel_settings.get("Camera Model", ""),
        "Lens Model": excel_settings.get("Lens Model", ""),
        "Illumination": excel_settings.get("Illumination", ""),
        "Viewing Angle": excel_settings.get("Viewing Angle", ""),
        "Shielding Gas": excel_settings.get("Shielding Gas", ""),
        "Current (A)": excel_settings.get("Current (A)", ""),
        "Voltage (V)": excel_settings.get("Voltage (V)", ""),
        "Heat Input": excel_settings.get("Heat Input (J/mm)", ""),
        "CTWD": excel_settings.get("CTWD (mm)", ""),
    }

    # Add measurements per class
    for feature, data in measurements.items():
        x_min = data.get("x_min")
        x_max = data.get("x_max")
        y_min = data.get("y_min")
        y_max = data.get("y_max")
        
        if x_min is not None and x_max is not None:
            data["x_avg"] = (x_min + x_max) / 2
        else:
            data["x_avg"] = None

        if y_min is not None and y_max is not None:
            data["y_avg"] = (y_min + y_max) / 2
        else:
            data["y_avg"] = None
    
    for feature in ["Arc Flash", "Solidification Pool", "Welding Wire"]:
        if feature in measurements:
            for key, val in measurements[feature].items():
                col = f"{feature}_{key}"
                row_data[col] = val

    # Initialize file if needed
    file_exists = os.path.exists(excel_path)
    if not file_exists:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Measurements"
        for i, key in enumerate(row_data.keys(), 1):
            cell = ws.cell(row=1, column=i, value=key)
            cell.font = Font(bold=True)
        wb.save(excel_path)

    # Append data
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    row = [row_data.get(h, "") for h in headers]
    ws.append(row)
    wb.save(excel_path)

def save_frame_images(raw_image, annotated_image, segmented_image, frame_counter):
    global experiment_folder
    subdirs = {
        "raw": "Raw Images",
        "annotated": "Annotated Images",
        "segmented": "Segmented Images"
    }

    # Create subdirectories inside the experiment folder
    for key, subdir in subdirs.items():
        os.makedirs(os.path.join(experiment_folder, subdir), exist_ok=True)

    if raw_image is not None:
        raw_path = os.path.join(experiment_folder, subdirs["raw"], f"frame_{frame_counter:05d}.png")
        cv2.imwrite(raw_path, raw_image)

    if annotated_image is not None:
        annotated_path = os.path.join(experiment_folder, subdirs["annotated"], f"frame_{frame_counter:05d}.png")
        cv2.imwrite(annotated_path, annotated_image)

    if segmented_image is not None:
        segmented_path = os.path.join(experiment_folder, subdirs["segmented"], f"frame_{frame_counter:05d}.png")
        cv2.imwrite(segmented_path, segmented_image)

def save_all_charts(graph_settings, canvases):
    if not graph_settings.get("save_charts", False):
        return  # Don't save charts unless specified
    
    save_folder = experiment_folder
    if not save_folder:
        print("No save_folder specified in graph_settings.")
        return

    os.makedirs(save_folder, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    for chart_name, canvas in canvases.items():
        fig = canvas.figure
        sanitized_name = chart_name.replace(" ", "_")
        filename = f"{sanitized_name}_{timestamp}.png"
        save_path = os.path.join(save_folder, filename)
        fig.savefig(save_path)
        print(f"Chart saved: {save_path}")

acquired_queue = queue.Queue()
save_image_queue = queue.Queue()
save_measurements_queue = queue.Queue()
draw_queue = queue.Queue()

def video_acquiring_worker():
    global is_recording
    while is_recording:
        raw_image, annotated_image, segmented_image, measurements, exposure_time = video_acquiring(recording_settings, annotation_settings)
        frame_data = {
            "frame_count": frame_count,
            "raw_image": raw_image,
            "annotated_image": annotated_image,
            "segmented_image": segmented_image,
            "measurements": measurements,
            "exposure_time": exposure_time
        }
        acquired_queue.put(frame_data)

def image_saving_worker():
    while is_recording:
        frame_data = save_image_queue.get()
        save_frame_images(
            frame_data["raw_image"],
            frame_data["annotated_image"],
            frame_data["segmented_image"],
            frame_data["frame_count"]
        )
        
def measurement_saving_worker():
    while is_recording:
        frame_data = save_measurements_queue.get()
        if raw_experiment_excel_data_settings.get("Save Raw Data in Excel") == True:
            save_measurements_to_excel(
                frame_data["frame_count"],
                frame_data["measurements"],
                raw_experiment_excel_data_settings,
                frame_data["exposure_time"]
            )

def draw_preview_worker():
    while is_recording:
        frame_data = draw_queue.get()
        window.after(0, draw_video_previews_from_frames, recording_settings, {
            "video_raw": frame_data["raw_image"],
            "video_annotated": frame_data["annotated_image"],
            "video_segmented": frame_data["segmented_image"]
        })
        
        

def update_recording_loop():
    if is_recording:
        try:
            if not acquired_queue.empty():
                frame_data = acquired_queue.get()

                # Dispatch to workers
                save_image_queue.put(frame_data)
                save_measurements_queue.put(frame_data)
                draw_queue.put(frame_data)

        finally:
            window.after(30, update_recording_loop)  # Schedule next GUI check

record_button = ctk.CTkButton(buttons_frame, 
                             text="Start Recording",
                             command=toggle_recording,
                             font=("Arial", 16, "bold"),
                             fg_color="green",
                             hover_color="dark green",
                             width=200,
                             height=50)
record_button.grid(row=0, column=6, padx=20, pady=10, sticky="ew")

def reset_recording():
    global start_time_integer, frame_count
    # Reset time tracking variables
    start_time_integer = 0
    frame_count = 0
    
    # Clear any existing preview frames
    for widget in image_frame.winfo_children():
        if isinstance(widget, ctk.CTkFrame):
            widget.destroy()
    
    # Clear our tracking dictionaries
    preview_frames.clear()
    preview_labels.clear()
    preview_texts.clear()
    
    # Update display with new empty previews based on current settings
    draw_video_previews_from_json_selection(recording_settings)

reset_button = ctk.CTkButton(buttons_frame,
                             text = "Reset Recording",
                             command = reset_recording,
                             font=("Arial", 16, "bold"),
                             fg_color="blue",
                             hover_color="dark blue",
                             width=200,
                             height=50)
reset_button.grid(row=0,column=7, padx=20, pady=10, sticky = "ew")


def update_graphs():
    if os.path.exists("graph_settings.json"):
        with open("graph_settings.json", 'r') as f:
            global graph_settings 
            graph_settings = json.load(f)
            render_charts()

def update_gui_from_settings():
    """Update GUI elements based on latest settings"""
    # Reload settings
    if os.path.exists("recording_settings.json"):
        with open("recording_settings.json", 'r') as f:
            global recording_settings
            recording_settings = json.load(f)
    
    # Add image viewing field to display image if the setting is selected
    reset_recording()
    draw_video_previews_from_json_selection(recording_settings)
    draw_image_previews_from_json_selection(recording_settings)

def update_class_values():
    if os.path.exists("class_values.json"):
        with open("class_values.json", 'r') as f:
            global class_values 
            class_values = json.load(f)
            render_charts()
    
# Add Image Viewing Field with fixed minimum size
image_frame = ctk.CTkFrame(master=window)
image_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=5)
image_frame.grid_columnconfigure(0, weight=1)
image_frame.grid_rowconfigure(0, weight=0)  # Title row doesn't expand
image_frame.grid_rowconfigure(1, weight=1)  # Content row expands
image_frame.grid_propagate(False)  # Prevent frame from shrinking
image_frame.configure(height=400)  # Set minimum height

image_title = ctk.CTkLabel(image_frame, text="Images", font=("Arial", 12, "bold"))
image_title.grid(row=0, column=0, sticky="w", padx=10, pady=5)

status_frame = ctk.CTkFrame(master=window)
status_frame.grid(row=2,column=0,sticky="nsew", padx=10, pady=10)
status_title = ctk.CTkLabel(status_frame, text="Recording Status:", font=("Arial", 12, "bold"))
status_title.grid(row=0,column=0,sticky="w", padx=10, pady=5)


camera_status_frame_title = ctk.CTkLabel(status_frame, text="Camera Status:", font=("Arial",12,"bold"))
camera_status_frame_title.grid(row=0,column=1,sticky="nsew",padx=10,pady=5)

# New big status display
big_status_display = ctk.CTkLabel(
    status_frame, 
    text="Idle",              # Example status text
    font=("Arial", 20, "bold"),       # Large font for emphasis
    text_color="yellow",               # Optional: use color to show status
    anchor="w"
)

def update_camera_status(label, state, detail=""):
    status_mapping = {
        'idle': ("Idle", "yellow"),
        'recording': ("Recording", "green"),
        'waiting_rsi': ("Waiting for RSI", "blue"),
        'tolerance_error': ("Out of Tolerance", "red"),
    }

    if state not in status_mapping:
        print(f"⚠️ Unknown status: {state}")
        return

    status_text, status_color = status_mapping[state]
    if detail:
        status_text += f"\n{detail}"

    # Only update if something changed
    if label.cget("text") != status_text or label.cget("text_color") != status_color:
        label.configure(text=status_text, text_color=status_color)



big_status_display.grid(row=1, column=0, sticky="ew", padx=10, pady=(5, 10))


big_camera_display = ctk.CTkLabel(status_frame,
                                  text="Disconnected",
                                  font=("Arial", 20, "bold"),
                                  text_color="yellow",
                                  anchor="e")
big_camera_display.grid(row=1,column=1, sticky="ew", padx=10, pady= (5,10))
# Optional: make sure the frame expands with window resizing
status_frame.grid_columnconfigure(0, weight=1)

def segment_raw_image(raw_image, segmentation_settings):
    if not segmentation_settings.get("apply_segmentation", False):
        return {}, raw_image

    # Load YOLOv8 model
    model_path = segmentation_settings["segmentation_file"]
    model = YOLO(model_path)

    # Perform segmentation
    results = model(raw_image)[0]

    # Exit early if nothing was detected
    if results.masks is None or results.boxes is None or len(results.boxes) == 0:
        return {}, raw_image

    # Desired classes
    target_classes = ["Welding Wire", "Solidification Zone", "Arc Flash"]
    class_names = ["Arc Flash", "Solidification Pool", "Welding Wire"]
    colors = {
        "Welding Wire": (255, 0, 0),
        "Solidification Zone": (0, 255, 0),
        "Arc Flash": (0, 0, 255),
    }

    class_data = {}
    overlay = raw_image.copy()

    found_any = False  # Flag to check if any relevant class was detected

    for mask, box, cls_id in zip(results.masks.data, results.boxes.data, results.boxes.cls):
        class_name = class_names[int(cls_id)]

        if class_name not in target_classes:
            continue

        found_any = True

        # Get bounding box
        x_min, y_min, x_max, y_max = map(int, box[:4])
        mask_np = mask.cpu().numpy()

        # Compute class area
        area = int(np.sum(mask_np))

        # Store class data
        class_data[class_name] = {
            "x_min": x_min,
            "x_max": x_max,
            "y_min": y_min,
            "y_max": y_max,
            "class_area": area
        }

        # Overlay colored mask
        if mask_np.shape[:2] != raw_image.shape[:2]:
            mask_np = cv2.resize(mask_np, (raw_image.shape[1], raw_image.shape[0]), interpolation=cv2.INTER_NEAREST)

        colored_mask = np.zeros_like(raw_image, dtype=np.uint8)
        colored_mask[mask_np.astype(bool)] = colors[class_name]
        overlay = cv2.addWeighted(overlay, 1.0, colored_mask, 0.5, 0)

    if not found_any:
        return {}, raw_image

    return class_data, overlay

def set_ET_time(recording_settings,time):
    if recording_settings.get("et_mode") == "Fixed":
        return recording_settings.get("et_fixed"), 0  # Fixed mode, loop count always 0

    elif recording_settings.get("et_mode") == "Iterate":
        et_start = recording_settings.get("et_start")
        et_end = recording_settings.get("et_end")
        et_step = recording_settings.get("et_step")
        time_increment = recording_settings.get("et_time")

        # Total number of unique steps before looping
        total_steps = ((int(et_end) - int(et_start)) // int(et_step)) + 1

        # How many time steps have passed
        steps_passed = int(time // int(time_increment))

        # Determine loop count and current step
        loop_count = steps_passed // total_steps
        current_step = steps_passed % total_steps

        # Calculate current exposure
        exposure_time = int(et_start) + int(current_step) * int(et_step)

        return exposure_time, loop_count

# def annotate_raw_image(frame, annotation_settings, exposure_time, loop_count, elapsed_time, fps):
#     annotated_frame = frame
#     frame_height, frame_width, _ = annotated_frame.shape

#     ### ─── TOP LEFT ───────────────────────────────────────────────
#     if annotation_settings["manual_settings"]["label_positions"].get("top-left"):
#         line_offset = 0
#         top_left_start_y = 30
#         spacing = 20

#         # Exposure Time
#         cv2.putText(annotated_frame, f"ET: {exposure_time}", (10, top_left_start_y + spacing * line_offset),
#                     cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
#         line_offset += 1

#         # Time
#         if annotation_settings["manual_settings"]["top_left_fields"]["Time"].get("show"):
#             text = round(elapsed_time,2)
#             cv2.putText(annotated_frame, f"Time: {text}", (10, top_left_start_y + spacing * line_offset),
#                         cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
#             line_offset += 1

#         # FA
#         if annotation_settings["manual_settings"]["top_left_fields"]["FA"].get("show"):
#             text = annotation_settings["manual_settings"]["top_left_fields"]["FA"].get("value")
#             cv2.putText(annotated_frame, f"FA: {text}", (10, top_left_start_y + spacing * line_offset),
#                         cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)

#     ### ─── TOP RIGHT ──────────────────────────────────────────────
#     if annotation_settings["manual_settings"]["label_positions"].get("top-right"):
#         line_offset = 0
#         top_right_start_y = 30
#         spacing = 20

#         fields = annotation_settings["manual_settings"]["top_right_fields"]
#         if fields["FPS"].get("show"):
#             text = str(round(fps,2))
#             cv2.putText(annotated_frame, "FPS: " + text,
#                         (frame_width - 235, top_right_start_y + spacing * line_offset ),
#                         cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
#             line_offset += 1
            
#         if fields["Running"].get("show"):
#             text = "Basler UI"
#             cv2.putText(annotated_frame, "Running: " + text,
#                         (frame_width - 235, top_right_start_y + spacing * line_offset ),
#                         cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
#             line_offset += 1
            
#         if fields["Output"].get("show"):
#             text = "Sample"
#             cv2.putText(annotated_frame, "Output: " + text,
#                         (frame_width - 235, top_right_start_y + spacing * line_offset),
#                         cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
#             line_offset += 1
            
#         for field in ["Illum", "Shielding Gas", "Note"]:
#             if fields[field].get("show"):
#                 text = f"{field}: {fields[field].get('value')}"
#                 text_size = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, 0.5, 1)[0]
#                 cv2.putText(annotated_frame, text,
#                             (frame_width - 235, top_right_start_y + spacing * line_offset),
#                             cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
#                 line_offset += 1

#     ### ─── BOTTOM LEFT ────────────────────────────────────────────
#     if annotation_settings["manual_settings"]["label_positions"].get("bottom-left"):
#         line_offset = 0
#         bottom_left_start_y = frame_height - 100
#         spacing = 20

#         fields = annotation_settings["manual_settings"]["bottom_left_fields"]
#         for field in ["Material", "Job Number", "Wire Feed Speed", "Travel Speed"]:
#             if fields[field].get("show"):
#                 text = f"{field}: {fields[field].get('value')}"
#                 cv2.putText(annotated_frame, text,
#                             (10, bottom_left_start_y + spacing * line_offset),
#                             cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
#                 line_offset += 1

#     ### ─── BOTTOM RIGHT ───────────────────────────────────────────
#     if annotation_settings["manual_settings"]["label_positions"].get("bottom-right"):
#         line_offset = 0
#         bottom_right_start = frame_height - 120
#         spacing = 20

#         # Camera
#         if annotation_settings["manual_settings"]["bottom_right_fields"]["Camera"].get("show"):
#             text = annotation_settings["manual_settings"]["bottom_right_fields"]["Camera"].get("value")
#             cv2.putText(annotated_frame, text,
#                         (frame_width - 200, bottom_right_start + spacing * line_offset),
#                         cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
#             line_offset += 1

#         # Lens + Viewing Angle
#         fields = annotation_settings["manual_settings"]["bottom_right_fields"]
#         show_lens = fields["Lens"].get("show")
#         show_angle = fields["Viewing Angle"].get("show")
#         lens_val = fields["Lens"].get("value")
#         angle_val = fields["Viewing Angle"].get("value")

#         if show_lens or show_angle:
#             combined = []
#             if show_lens:
#                 combined.append(lens_val)
#             if show_angle:
#                 combined.append(angle_val)
#             text = " | ".join(combined)
#             cv2.putText(annotated_frame, text,
#                         (frame_width - 200, bottom_right_start + spacing * line_offset),
#                         cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
#             line_offset += 1

#         # Focus
#         if fields["Focus"].get("show") & fields["Aperature"].get("show"):
#             cv2.putText(annotated_frame, "F:" + fields["Focus"].get("value") + " | A:" + fields["Aperature"].get("value"),
#                         (frame_width - 200, bottom_right_start + spacing * line_offset),
#                         cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
#             line_offset += 1
#         elif fields["Focus"].get("show"):
#             cv2.putText(annotated_frame, "F:" + fields["Focus"].get("value"),
#                         (frame_width - 200, bottom_right_start + spacing * line_offset),
#                         cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
#             line_offset += 1

#         # Aperture (note: JSON typo was "Aperature" but use "Aperture" for consistency)
        
            
#         elif fields["Aperature"].get("show"):
#             cv2.putText(annotated_frame, "A:" + fields["Aperature"].get("value"),
#                         (frame_width - 200, bottom_right_start + spacing * line_offset),
#                         cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
#             line_offset += 1

#         # Distance
#         if fields["Distance"].get("show"):
#             text = f"Distance: {fields['Distance'].get('value')}"
#             cv2.putText(annotated_frame, text,
#                         (frame_width - 200, bottom_right_start + spacing * line_offset),
#                         cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
#             line_offset += 1

#         # CTWD
#         if fields["CTWD (mm)"].get("show"):
#             text = f"dNtW (mm): {fields['CTWD (mm)'].get('value')}"
#             cv2.putText(annotated_frame, text,
#                         (frame_width - 200, bottom_right_start + spacing * line_offset),
#                         cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)

#     return annotated_frame

def calculate_fps(frame_count, start_time):
    # Calculate the fps of the video based on the current time, start time, and frames currently collected.
    elapsed_time = time.time() - start_time
    fps = frame_count / elapsed_time if elapsed_time > 0 else 0
    return fps, elapsed_time

def calculate_instantaneous_fps(prev_time):
    current_time = time.time()
    time_diff = current_time - prev_time
    fps = 1.0 / time_diff if time_diff > 0 else 0
    return fps, current_time

def parse_rsi_data(xml_data):
    """Parses RSI XML data and extracts the CAM value."""
    try:
        root = ET.fromstring(xml_data)
        cam_value = int(root.find("CAM").text)  # Convert CAM to integer (0 or 1)
        return cam_value
    except Exception as e:
        print(f"Error parsing XML: {e}")
        return None

def receive_data():
    """Receives RSI data, detects CAM transition, and updates the global `cam` variable."""
    global cam
    with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
        s.bind((R_IP, R_PORT))
        print(f"Listening on {R_IP}:{R_PORT}")
        
        while True:
            try:
                data, addr = s.recvfrom(1024)
                data_str = data.decode('utf-8')
                
                new_cam = parse_rsi_data(data_str)
                if new_cam is not None:
                    with cam_lock:
                        cam = new_cam
                else:
                    with cam_lock:
                        cam = 0
            except Exception as e:
                print(f"Error receiving data: {e}")
x_max_cumulative = {
    "Arc Flash": [],
    "Solidification Pool": [],
    "Welding Wire": []
}

x_average_cumulative = {
    "Arc Flash": [],
    "Solidification Pool": [],
    "Welding Wire": []
}

x_min_cumulative = {
    "Arc Flash": [],
    "Solidification Pool": [],
    "Welding Wire": []
}

y_max_cumulative = {
    "Arc Flash": [],
    "Solidification Pool": [],
    "Welding Wire": []
}

y_average_cumulative = {
    "Arc Flash": [],
    "Solidification Pool": [],
    "Welding Wire": []
}

y_min_cumulative = {
    "Arc Flash": [],
    "Solidification Pool": [],
    "Welding Wire": []
}

area_average_cumulative = {
    "Arc Flash": [],
    "Solidification Pool": [],
    "Welding Wire": []
}
area_std_deviation_cumulative = {
    "Arc Flash": [],
    "Solidification Pool": [],
    "Welding Wire": []
}
y_avg_std_deviation_cumulative = {
    "Arc Flash": [],
    "Solidification Pool": [],
    "Welding Wire": []
}
x_avg_std_deviation_cumulative = {
    "Arc Flash": [],
    "Solidification Pool": [],
    "Welding Wire": []
}

frame_counter = 0  # Persistent frame counter

output_data = {
    "X Average": {
        "Arc Flash": [],
        "Solidification Pool": [],
        "Welding Wire": [],
        "Graph Frame Index": []
    },
    "X Maximum": {
        "Arc Flash": [],
        "Solidification Pool": [],
        "Welding Wire": [],
        "Graph Frame Index": []
    }
}

# Global metric configuration
class_names = ["Arc Flash", "Solidification Pool", "Welding Wire"]
metric_configs = {
    "X Average": {"fields": ("x_min", "x_max"), "func": lambda x1, x2: (x1 + x2) / 2, "type": "average"},
    "X Maximum": {"fields": ("x_max",), "func": lambda x: x, "type": "average"},
    "X Minimum": {"fields": ("x_min",), "func": lambda x: x, "type": "average"},
    "Y Average": {"fields": ("y_min", "y_max"), "func": lambda y1, y2: (y1 + y2) / 2, "type": "average"},
    "Y Maximum": {"fields": ("y_max",), "func": lambda y: y, "type": "average"},
    "Y Minimum": {"fields": ("y_min",), "func": lambda y: y, "type": "average"},

    "Class Area": {"fields": ("area",), "func": lambda a: a, "type": "average"},
    "Class Area Standard Deviation": {"fields": ("area",), "func": np.std, "type": "std"},
    "Y Average Standard Deviation": {"fields": ("y_average",), "func": np.std, "type": "std"},
    "X Average Standard Deviation": {"fields": ("x_average",), "func": np.std, "type": "std"},
}

# Initialize cumulative data structures
cumulative_data = {
    metric: {class_name: [] for class_name in class_names}
    for metric in metric_configs
}

def compute_and_append(metric, class_name, values, cumulative_data):
    if None in values:
        return

    metric_type = metric_configs[metric]["type"]

    if metric_type == "average":
        func = metric_configs[metric]["func"]
        result = func(*values) if len(values) > 1 else func(values[0])
        cumulative_data[metric][class_name].append(result)
    
    elif metric_type == "std":
        # For std metrics, just store the raw input value(s) for now
        cumulative_data[metric][class_name].extend(values)

def should_output(metric, frame_counter, frame_count_settings):
    count = frame_count_settings.get(metric, 10)
    return frame_counter != 0 and frame_counter % int(count) == 0

def output_metric_data(metric, frame_counter, cumulative_data, output_data):
    if metric not in output_data:
        output_data[metric] = {name: [] for name in class_names}
        output_data[metric]["Graph Frame Index"] = []

    metric_type = metric_configs[metric]["type"]
    func = metric_configs[metric]["func"]

    for class_name in class_names:
        values = cumulative_data[metric][class_name]
        
        if not values:
            output = 0
        elif metric_type == "average":
            output = np.mean(values)
        elif metric_type == "std":
            output = np.std(values)
        else:
            output = 0

        output_data[metric][class_name].append(output)

    output_data[metric]["Graph Frame Index"].append(frame_counter)

    # Reset for next batch
    for class_name in class_names:
        cumulative_data[metric][class_name].clear()

def calculate_metrics_over_frames(all_class_data, settings, output_data):
    global frame_counter
    graph_key = []
    metrics_settings = settings.get("metrics", {})
    frame_count_settings = settings.get("frame_counts", {})

    for metric, config in metric_configs.items():
        if not metrics_settings.get(metric):
            continue

        for class_name in class_names:
            class_data = all_class_data[class_name]
            values = [class_data.get(field) for field in config["fields"]]

            if all(v is not None for v in values):
                compute_and_append(metric, class_name, values, cumulative_data)

        if should_output(metric, frame_counter, frame_count_settings):
            output_metric_data(metric, frame_counter, cumulative_data, output_data)
            graph_key.append(metric)
    frame_counter = frame_counter +1
    return graph_key

def create_visualization(frame, masks, boxes, measurements):
    """Create visualization with masks and measurements."""
    output = frame.copy()
    class_names = ['Arc Flash', 'Solidification Pool', 'Welding Wire']
    colors = {
        0: (255, 0, 0),    # Arc Flash - Blue
        1: (0, 255, 0),    # Solidification Pool - Green
        2: (0, 0, 255)     # Welding Wire - Red
    }

    for cls in range(len(class_names)):
        mask = masks[cls]
        if mask is not None and isinstance(mask, np.ndarray) and mask.ndim == 2:
            # Apply color mask
            colored_mask = np.zeros_like(frame)
            colored_mask[mask > 0] = colors[cls]
            output = cv2.addWeighted(output, 0.7, colored_mask, 0.3, 0)

            # Add label with measurements
            class_name = class_names[cls]
            measurement = measurements.get(class_name, {})
            
            if measurement.get('width') is not None:
                area = measurement['area']
                width = measurement['width']
                height = measurement['height']

                # Unpack boxes[cls] and ignore extra elements
                box = boxes.get(cls)
                if box is not None:
                    # Extract only the first four values: x1, y1, x2, y2
                    x1, y1, x2, y2 = box[:4].cpu().numpy()  # Convert tensor to numpy array (if it's a tensor)
                else:
                    # Default to the measurement's min coordinates if box is None
                    x1, y1 = measurement.get('x_min', 0), measurement.get('y_min', 0)
                    x2, y2 = x1 + measurement.get('width', 0), y1 + measurement.get('height', 0)

                label = f"{class_name} ({width}x{height}, {area:,d}px)"
                cv2.putText(output, label, (int(x1), int(y1) - 10), 
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 2)
        
    return output

model = YOLO(segmentation_settings["segmentation_file"]) 
class_names = ['Arc Flash', 'Solidification Pool', 'Welding Wire']
device = "cuda" if torch.cuda.is_available() else "cpu"

def process_frame(frame, model, class_names):
    """Optimized: Process a frame and return segmentation results with timing."""
    # t_start = time.time()

    orig_H, orig_W, _ = frame.shape
    

    # model.to(device)
    # model.eval()

    # Convert frame to tensor
    # t_preprocessing_start = time.time()
    input_size = 640
    resized_frame = cv2.resize(frame, (input_size, input_size))

    frame_tensor = (
        torch.from_numpy(resized_frame)
        .permute(2, 0, 1)
        .unsqueeze(0)
        .float()
        / 255.0
    )

    if device == "cuda":
        frame_tensor = frame_tensor.to(device).half()
        model = model.half()
    else:
        frame_tensor = frame_tensor.to(device)

    # t_inference_start = time.time()
    results = model(frame_tensor)[0]
    # t_post_inference_start = time.time()

    # Initialize data structures
    data = {name: {
        'area': None,
        'x_min': None,
        'x_max': None,
        'y_min': None,
        'y_max': None,
        'width': None,
        'height': None
    } for name in class_names}

    largest_masks = {i: None for i in range(len(class_names))}
    largest_boxes = {i: None for i in range(len(class_names))}
    largest_areas = {i: 0 for i in range(len(class_names))}

    if hasattr(results, 'masks') and results.masks is not None:
        masks = results.masks.data
        boxes = results.boxes.data

        masks = F.interpolate(masks.unsqueeze(1).float(), size=(orig_H, orig_W), mode='nearest').squeeze(1)

        for i, (mask, box) in enumerate(zip(masks, boxes)):
            cls = int(box[5])
            cls_name = class_names[cls]

            binary_mask = (mask > 0.5)
            area = binary_mask.sum().item()

            if area > largest_areas[cls]:
                largest_areas[cls] = area
                largest_masks[cls] = binary_mask.detach().cpu().numpy().astype(np.uint8)
                largest_boxes[cls] = box.detach().cpu()

                ys, xs = torch.where(binary_mask)
                if ys.numel() > 0:
                    x_min = int(xs.min().item())
                    x_max = int(xs.max().item())
                    y_min = int(ys.min().item())
                    y_max = int(ys.max().item())
                    width = x_max - x_min
                    height = y_max - y_min

                    data[cls_name].update({
                        'area': int(area),
                        'x_min': x_min,
                        'x_max': x_max,
                        'y_min': y_min,
                        'y_max': y_max,
                        'width': width,
                        'height': height
                    })

    # t_end = time.time()

    # Print timing for diagnostics
    # print(f"Total time: {t_end - t_start:.4f}s")
    # print(f" - Preprocessing: {t_inference_start - t_preprocessing_start:.4f}s")
    # print(f" - Inference: {t_post_inference_start - t_inference_start:.4f}s")
    # print(f" - Post-processing: {t_end - t_post_inference_start:.4f}s")

    return data, largest_masks, largest_boxes


global metric_to_class_key
metric_to_class_key = {
    "X Average": "x_average",
    "X Maximum": "x_max",
    "X Minimum": "x_min",
    "Y Average": "y_average",
    "Y Maximum": "y_max",
    "Y Minimum": "y_min",
    "X Average Standard Deviation" : "x_avg_std_deviation",
    "Y Average Standard Deviation" : "y_avg_std_deviation",
    "Class Area" : "area_average",
    "Class Area Standard Deviation" : "area_std_deviation"
}
device = "cuda" if torch.cuda.is_available() else "cpu"
model = YOLO("yolov8n-seg.pt")  # Or however you load your model
model.to(device)
model.eval()
class_names = ["Arc Flash", "Solidification Pool", "Welding Wire"]  # Or your own list like ["Wire", "Arc", ...]

def initialize_camera_settings(camera):
    try:
        # Basic settings that should work on most Basler cameras
        if camera.GetDeviceInfo().IsGigEDevice():
            # GigE specific settings
            if hasattr(camera, 'GevSCPSPacketSize'):
                camera.GevSCPSPacketSize.SetValue(9000)
            if hasattr(camera, 'GevSCPD'):
                camera.GevSCPD.SetValue(0)
            
        # Check and set buffer and queue settings
        if hasattr(camera, 'MaxNumBuffer'):
            camera.MaxNumBuffer.SetValue(5)
        if hasattr(camera, 'OutputQueueSize'):
            camera.OutputQueueSize.SetValue(1)
            
        # Set acquisition settings if available
        if hasattr(camera, 'GrabStrategy'):
            camera.GrabStrategy.SetValue('LatestImageOnly')
        if hasattr(camera, 'AcquisitionFrameRateEnable'):
            camera.AcquisitionFrameRateEnable.SetValue(True)
            if hasattr(camera, 'AcquisitionFrameRate'):
                camera.AcquisitionFrameRate.SetValue(30.0)
                
        camera._settings_initialized = True
        print("✅ Camera settings initialized successfully")
        
    except Exception as e:
        print(f"❌ Error initializing camera settings: {e}")
        # Set flag even if failed to prevent repeated attempts
        camera._settings_initialized = True



def video_acquiring(recording_settings, annotation_settings):
    import time  # Make sure this is imported
    global frame_count, start_time_integer, start_time, prev_frame_time
    global segmentation_settings, image_paths, output_data, metric_to_class_key

    t_start = time.perf_counter()

    if recording_settings.get("rsi_mode") == "Automatic":
        update_camera_status(big_status_display,"waiting_rsi")
        cam_value = receive_data()
        while cam_value != 1:
            time.sleep(0.05)
            cam_value = receive_data()
        update_camera_status(big_status_display,"recording")
    else:
        update_camera_status(big_status_display,"recording")
    t_cam_check = time.perf_counter()

    if start_time_integer == 0:
        start_time = time.time()
        prev_frame_time = start_time
        start_time_integer = 1

    elapsed_time = time.time() - start_time
    frame_count += 1

    exposure_time, loop_count = set_ET_time(recording_settings, elapsed_time)
    camera.ExposureTimeRaw.SetValue(int(exposure_time))
    t_exposure = time.perf_counter()
    
    
    
    
    grab_result = camera.GrabOne(1000)  # Keep timeout but with optimized settings
    
    t_grab = time.perf_counter()

    raw_image = None
    annotated_image = None
    segmented_image = None
    measurements = {}

    if grab_result and grab_result.IsValid() and grab_result.GrabSucceeded():
        img_array = grab_result.Array
        raw_rgb = cv2.cvtColor(img_array, cv2.COLOR_BAYER_BG2RGB)
        fps, prev_frame_time = calculate_instantaneous_fps(prev_frame_time)

        if recording_settings.get("video_raw"):
            raw_image = raw_rgb.copy()

        if recording_settings.get("video_annotated"):
            annotated_image = annotate_raw_image(
                raw_rgb.copy(), annotation_settings, exposure_time,
                loop_count, elapsed_time, fps
            )

        t_annotation = time.perf_counter()

        if recording_settings.get("video_segmented"):
            t0 = time.perf_counter()
            # image_path = image_paths["image_raw"]
            # t1 = time.perf_counter()
            # print(f"[Timing] Fetch image path: {(t1 - t0)*1000:.2f} ms")

            # image = Image.open(image_path).convert('RGB')
            # t2 = time.perf_counter()
            # print(f"[Timing] Load and convert image: {(t2 - t1)*1000:.2f} ms")

            # sample_image = np.array(image, dtype=np.uint8)
            # t3 = time.perf_counter()
            # print(f"[Timing] Convert image to NumPy array: {(t3 - t2)*1000:.2f} ms")

            measurements, largest_masks, largest_boxes = process_frame(raw_image, model, class_names)
            t4 = time.perf_counter()
            print(f"[Timing] Run segmentation (process_frame): {(t4 - t3)*1000:.2f} ms")

            segmented_image = create_visualization(raw_image, largest_masks, largest_boxes, measurements)
            t5 = time.perf_counter()
            print(f"[Timing] Create visualization: {(t5 - t4)*1000:.2f} ms")

            total_time = t5 - t0
            print(f"[Timing] Total time for video_segmented block: {total_time*1000:.2f} ms")

        t_segmentation = time.perf_counter()

        if recording_settings.get("rsi_mode") == "Automatic":
            cam_value = receive_data()
            if cam_value == 0:
                print("📴 RSI signal ended — stopping recording.")
                toggle_recording()
    else:
        print("❌ Grab failed or returned invalid result.")

    t_rsi = time.perf_counter()
    if recording_settings.get("video_segmented"):
        result = calculate_metrics_over_frames(measurements, graph_settings, output_data)
        if result:
            for key in result:
                ax = axes[key]
                metric = metric_to_class_key[key]
                for feature in ["Arc Flash", "Solidification Pool", "Welding Wire"]:
                    x_data = output_data[key]["Graph Frame Index"]
                    y_data = output_data[key][feature]

                    if feature == "Solidification Pool":
                        feature = "Solidification Zone"

                    desired_value = int(class_values[str(material_selection)][feature][metric]["value"])
                    tol_pos = int(class_values[str(material_selection)][feature][metric]["pos_tolerance"])
                    tol_neg = int(class_values[str(material_selection)][feature][metric]["neg_tolerance"])

                    if segmentation_settings.get("compare_values"):
                        for i, y in enumerate(y_data):
                            if y > desired_value + tol_pos or y < desired_value - tol_neg:
                                detail = (
                                    f"{feature} - {metric} out of tolerance at frame {i} "
                                    f"(Value: {y}, Expected: {desired_value} ±{tol_neg}/{tol_pos})"
                                )
                                print(f"[Tolerance Violation] {detail}")
                                update_camera_status(big_status_display,"tolerance_error", detail)

                    plot_feature_on_axes(
                        ax, feature, x_data, y_data,
                        desired_value, tol_pos, tol_neg, key
                    )
                canvases[key].draw()

    t_plot = time.perf_counter()

    # --- Print step durations ---
    print(f"[Timing] RSI check:              {(t_cam_check - t_start):.4f} s")
    print(f"[Timing] Exposure config:        {(t_exposure - t_cam_check):.4f} s")
    print(f"[Timing] Frame grab:             {(t_grab - t_exposure):.4f} s")
    print(f"[Timing] Annotation:             {(t_annotation - t_grab):.4f} s")
    print(f"[Timing] Segmentation + masks:   {(t_segmentation - t_annotation):.4f} s")
    print(f"[Timing] RSI post check:         {(t_rsi - t_segmentation):.4f} s")
    print(f"[Timing] Metric calc + plotting: {(t_plot - t_rsi):.4f} s")
    print(f"[Timing] Total video_acquiring:  {(t_plot - t_start):.4f} s\n")

    return raw_image, annotated_image, segmented_image, measurements, exposure_time



def establish_preview_frames(recording_settings):
    """Create persistent preview frames once"""
    preview_row = 1
    preview_col = 0
    i = 1
    
    for key in ["video_raw", "video_annotated", "video_segmented"]:
        if recording_settings.get(key):
            if key not in preview_frames:
                image_frame.grid_columnconfigure(i, weight=1)
                
                # Create frame
                preview_frames[key] = ctk.CTkFrame(master=image_frame)
                preview_frames[key].grid(row=preview_row, column=preview_col, padx=10, pady=10)
                
                # Create label for image
                preview_labels[key] = ctk.CTkLabel(master=preview_frames[key], text="")
                preview_labels[key].pack()
                
                # Create text label
                text = key.replace("_", " ").title()
                preview_texts[key] = ctk.CTkLabel(preview_frames[key], text=text)
                preview_texts[key].pack(pady=(5, 0))
                
                preview_col += 1
                i += 1

def draw_video_previews_from_frames(recording_settings, frame_dict):
    """Update existing preview frames with new images, resized based on number of active previews."""
    if not preview_frames:
        establish_preview_frames(recording_settings)
    
    # Determine how many previews are being shown
    
    active_keys = [key for key in ["video_raw", "video_annotated", "video_segmented"]
                   if recording_settings.get(key) and key in frame_dict]
    
    num_previews = len(active_keys)
    if num_previews == 0:
        return

    # Base total preview area (you can adjust these)
    max_width = 2000
    max_height = 1000

    # Calculate optimal preview size
    preview_width = max_width // num_previews
    preview_height = int(max_height * 0.5)  # adjust height proportionally

    for key in active_keys:
        try:
            frame = frame_dict[key]
            if frame is not None:
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                img_pil = Image.fromarray(frame_rgb)

                img_pil = img_pil.resize((preview_width, preview_height))

                # Update existing label with resized image
                ctk_img = ctk.CTkImage(
                    light_image=img_pil,
                    dark_image=img_pil,
                    size=(preview_width, preview_height)
                )
                preview_labels[key].configure(image=ctk_img)
                preview_labels[key].image = ctk_img  # Prevent garbage collection

        except Exception as e:
            print(f"Error updating preview for {key}:", e)

def draw_video_previews_from_json_selection(recording_settings):
    # Add image viewing field to display image if the setting is selected in the submenu.
    preview_row = 1
    preview_col = 0
    i=1
    
    for key in ["video_raw", "video_annotated", "video_segmented"]:
        if recording_settings.get(key):
            try:
                image_frame.grid_columnconfigure(i, weight=1)
                # Load and resize image with proper parameters
                img = Image.open(image_paths[key])
                # img = img.resize((650, 450), Image.Resampling.LANCZOS)  # Fixed resize parameters
                photo_img = ImageTk.PhotoImage(img)

                # Create a sub-frame to hold image + label
                preview_frame = ctk.CTkFrame(master=image_frame)
                preview_frame.grid(row=preview_row, column=preview_col, padx=10, pady=10)

                
                # Image label (no text)
                # photo_img = ctk.CTkImage(light_image=img, size=(550,400))  # auto-resizes

                # img_label = ctk.CTkLabel(master=preview_frame, image=photo_img, text="")
                # img_label.pack()

                # Text label underneath
                text = key.replace("_", " ").title()
                text_label = ctk.CTkLabel(master=preview_frame, text=text)
                text_label.pack(pady=(5, 0))

                preview_col += 1
                i += 1 
            except Exception as e:
                print(f"Error loading {key} preview:", e)
                
def draw_image_previews_from_json_selection(recording_settings):
    # Add image viewing field to display image if the setting is selected in the submenu.
    preview_row = 1
    preview_col = 0
    i=1
    
    for key in ["image_raw", "image_annotated", "image_segmented"]:
        if recording_settings.get(key):
            try:
                image_frame.grid_columnconfigure(i, weight=1)
                # Load and resize image with proper parameters
                img = Image.open(image_paths[key])
                # img = img.resize((650, 450), Image.Resampling.LANCZOS)  # Fixed resize parameters
                photo_img = ImageTk.PhotoImage(img)

                # Create a sub-frame to hold image + label
                preview_frame = ctk.CTkFrame(master=image_frame)
                preview_frame.grid(row=preview_row, column=preview_col, padx=10, pady=10)

                # Image label (no text)
                photo_img = ctk.CTkImage(light_image=img, size=(550,400))  # auto-resizes

                img_label = ctk.CTkLabel(master=preview_frame, image=photo_img, text="")
                img_label.pack()

                # Text label underneath
                text = key.replace("_", " ").title()
                text_label = ctk.CTkLabel(master=preview_frame, text=text)
                text_label.pack(pady=(5, 0))

                preview_col += 1
                i += 1 
            except Exception as e:
                print(f"Error loading {key} preview:", e)
                
draw_video_previews_from_json_selection(recording_settings)
draw_image_previews_from_json_selection(recording_settings)




def plot_feature_on_axes(ax, feature_name, x_data, y_data,
                         desired_value=None, tol_pos=None, tol_neg=None,
                         graph_key=None):
    FEATURE_COLORS = {
        "Solidification Zone": "#1f77b4",
        "Welding Wire": "#ff7f0e",
        "Arc Flash": "#2ca02c"
    }

    color = FEATURE_COLORS.get(feature_name, "black")

    if graph_key not in line_refs:
        line_refs[graph_key] = {}
    if graph_key not in fill_refs:
        fill_refs[graph_key] = {}
    if graph_key not in centerline_refs:
        centerline_refs[graph_key] = {}

    # Plot or update main line
    if feature_name not in line_refs[graph_key]:
        line, = ax.plot(x_data, y_data, label=feature_name, color=color)
        line_refs[graph_key][feature_name] = line
    else:
        line = line_refs[graph_key][feature_name]
        line.set_xdata(x_data)
        line.set_ydata(y_data)

    x_min = min(x_data)
    x_max = max(x_data)

    # Plot or update shaded tolerance band
    if desired_value is not None and tol_pos is not None and tol_neg is not None:
        lower_bound = desired_value - tol_neg
        upper_bound = desired_value + tol_pos

        if feature_name in fill_refs[graph_key]:
            fill = fill_refs[graph_key][feature_name]
            fill.remove()  # Reuse logic can go here instead
        fill = ax.axhspan(lower_bound, upper_bound, color=color, alpha=0.2)
        fill_refs[graph_key][feature_name] = fill

        # Update or create centerline
        if feature_name in centerline_refs[graph_key]:
            centerline = centerline_refs[graph_key][feature_name]
            centerline.set_ydata([desired_value, desired_value])
        else:
            centerline = ax.axhline(desired_value, color=color, linestyle='--', linewidth=1)
            centerline_refs[graph_key][feature_name] = centerline

    if x_min == x_max:
        x_min -= 0.5
        x_max += 0.5
    ax.set_xlim(left=x_min, right=x_max)

    ax.relim()
    ax.autoscale_view(scalex=False, scaley=True)

def render_charts():
    # Clear the frame first
    for widget in chart_frame.winfo_children():
        widget.destroy()
    fig_w=3.5
    fig_h=3.5

    # Draw frames for each graph type
    
    show_charts = graph_settings.get("show_charts", {})
    chart_metrics = graph_settings.get("metrics",{})
    
    # Define tolerance and good values
    material_selection 
    data=class_values[material_selection]
    global axes
    axes = {}
    global canvases 
    canvases = {}
    if show_charts.get("X Position Values"):
        column_index=0
        x_position_frame = ctk.CTkFrame(chart_frame)
        x_position_frame_label = ctk.CTkLabel(x_position_frame, text= "X Position Values", font=("Arial", 12, "bold"))
        x_position_frame.grid(row=0, column=column_index, padx=20, pady=20)
        x_position_frame_label.grid(row=0, column=0, sticky="w", padx=10, pady=5)
        if chart_metrics.get("X Average"):
            fig = Figure(figsize=(fig_w, fig_h), dpi=100)
            axes["X Average"] = fig.add_subplot(111)
            ax = axes["X Average"]
            x_average_frame = ctk.CTkFrame(x_position_frame)
            x_average_frame.grid(row=1, column=column_index, sticky="w", padx=10, pady=5)

            # Set title and labels ONCE
            ax.set_title("X Average")
            ax.set_xlabel("Frame Index")
            ax.set_ylabel("Pixel Value")

            # Now call plot_feature_on_axes without title/xlabel/ylabel every time
            plot_feature_on_axes(ax, "Solidification Zone",
                                x_data=[0, 1, 2], y_data=[0, 1, 2],
                                desired_value=int(data['Solidification Zone']['x_average']['value']),
                                tol_pos=int(data['Solidification Zone']['x_average']['pos_tolerance']),
                                tol_neg=int(data['Solidification Zone']['x_average']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Welding Wire",
                                x_data=[0, 1, 2], y_data=[0, 2, 4],
                                desired_value=int(data['Welding Wire']['x_average']['value']),
                                tol_pos=int(data['Welding Wire']['x_average']['pos_tolerance']),
                                tol_neg=int(data['Welding Wire']['x_average']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Arc Flash",
                                x_data=[0, 1, 2], y_data=[0, 3, 6],
                                desired_value=int(data['Arc Flash']['x_average']['value']),
                                tol_pos=int(data['Arc Flash']['x_average']['pos_tolerance']),
                                tol_neg=int(data['Arc Flash']['x_average']['neg_tolerance']))

            canvas = FigureCanvasTkAgg(fig, master=x_average_frame)
            canvases["X Average"] = canvas  # <--- SAVE IT
            canvas.draw()
            canvas.get_tk_widget().grid(row=1, column=column_index)
            column_index += 1
        if chart_metrics.get("X Maximum"):
            fig=Figure(figsize=(fig_w,fig_h),dpi=100)
            axes["X Maximum"] = fig.add_subplot(111)
            ax=axes["X Maximum"]
            x_maximum_frame = ctk.CTkFrame(x_position_frame)
            x_maximum_frame.grid(row=1,column=column_index, sticky = "w", padx=10, pady=5)
            
            # Set title and labels ONCE
            ax.set_title("X Maximum")
            ax.set_xlabel("Frame Index")
            ax.set_ylabel("Pixel Value")

            
            # Now call plot_feature_on_axes without title/xlabel/ylabel every time
            plot_feature_on_axes(ax, "Solidification Zone",
                                x_data=[0, 1, 2], y_data=[0, 1, 2],
                                desired_value=int(data['Solidification Zone']['x_max']['value']),
                                tol_pos=int(data['Solidification Zone']['x_max']['pos_tolerance']),
                                tol_neg=int(data['Solidification Zone']['x_max']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Welding Wire",
                                x_data=[0, 1, 2], y_data=[0, 2, 4],
                                desired_value=int(data['Welding Wire']['x_max']['value']),
                                tol_pos=int(data['Welding Wire']['x_max']['pos_tolerance']),
                                tol_neg=int(data['Welding Wire']['x_max']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Arc Flash",
                                x_data=[0, 1, 2], y_data=[0, 3, 6],
                                desired_value=int(data['Arc Flash']['x_max']['value']),
                                tol_pos=int(data['Arc Flash']['x_max']['pos_tolerance']),
                                tol_neg=int(data['Arc Flash']['x_max']['neg_tolerance']))
            
            canvas = FigureCanvasTkAgg(fig, master=x_maximum_frame)
            canvases["X Maximum"] = canvas  # <--- SAVE IT
            canvas.draw()
            canvas.get_tk_widget().grid(row=1,column=column_index)
            column_index=column_index+1
            
        if chart_metrics.get("X Minimum"):
            fig=Figure(figsize=(fig_w,fig_h),dpi=100)
            axes["X Minimum"] = fig.add_subplot(111)
            ax=axes["X Minimum"]
            x_minimum_frame = ctk.CTkFrame(x_position_frame)
            x_minimum_frame.grid(row=1,column=column_index, sticky = "w", padx=10, pady=5)
            # Set title and labels ONCE
            ax.set_title("X Minimum")
            ax.set_xlabel("Frame Index")
            ax.set_ylabel("Pixel Value")

            
            # Now call plot_feature_on_axes without title/xlabel/ylabel every time
            plot_feature_on_axes(ax, "Solidification Zone",
                                x_data=[0, 1, 2], y_data=[0, 1, 2],
                                desired_value=int(data['Solidification Zone']['x_min']['value']),
                                tol_pos=int(data['Solidification Zone']['x_min']['pos_tolerance']),
                                tol_neg=int(data['Solidification Zone']['x_min']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Welding Wire",
                                x_data=[0, 1, 2], y_data=[0, 2, 4],
                                desired_value=int(data['Welding Wire']['x_min']['value']),
                                tol_pos=int(data['Welding Wire']['x_min']['pos_tolerance']),
                                tol_neg=int(data['Welding Wire']['x_min']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Arc Flash",
                                x_data=[0, 1, 2], y_data=[0, 3, 6],
                                desired_value=int(data['Arc Flash']['x_min']['value']),
                                tol_pos=int(data['Arc Flash']['x_min']['pos_tolerance']),
                                tol_neg=int(data['Arc Flash']['x_min']['neg_tolerance']))
            
            
            canvas = FigureCanvasTkAgg(fig, master=x_minimum_frame)
            canvases["X Minimum"] = canvas  # <--- SAVE IT
            canvas.draw()
            canvas.get_tk_widget().grid(row=1,column=column_index)
            column_index=+1
            
            
    if show_charts.get("Y Position Values"):
        column_index=0
        y_position_frame = ctk.CTkFrame(chart_frame)
        y_position_frame_label = ctk.CTkLabel(y_position_frame, text= "Y Position Values", font=("Arial", 12, "bold"))
        y_position_frame.grid(row=1, column=0, padx=20, pady=20)
        y_position_frame_label.grid(row=0, column=0, sticky="w", padx=10, pady=5)
        
        if chart_metrics.get("Y Average"):
            fig=Figure(figsize=(fig_w,fig_h),dpi=100)
            axes["Y Average"] = fig.add_subplot(111)
            ax=axes["Y Average"]
            y_average_frame = ctk.CTkFrame(y_position_frame)
            y_average_frame.grid(row=1,column=column_index, sticky = "w", padx=10, pady=5)
            # Set title and labels ONCE
            ax.set_title("Y Average")
            ax.set_xlabel("Frame Index")
            ax.set_ylabel("Pixel Value")

            
            # Now call plot_feature_on_axes without title/xlabel/ylabel every time
            plot_feature_on_axes(ax, "Solidification Zone",
                                x_data=[0, 1, 2], y_data=[0, 1, 2],
                                desired_value=int(data['Solidification Zone']['y_average']['value']),
                                tol_pos=int(data['Solidification Zone']['y_average']['pos_tolerance']),
                                tol_neg=int(data['Solidification Zone']['y_average']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Welding Wire",
                                x_data=[0, 1, 2], y_data=[0, 2, 4],
                                desired_value=int(data['Welding Wire']['y_average']['value']),
                                tol_pos=int(data['Welding Wire']['y_average']['pos_tolerance']),
                                tol_neg=int(data['Welding Wire']['y_average']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Arc Flash",
                                x_data=[0, 1, 2], y_data=[0, 3, 6],
                                desired_value=int(data['Arc Flash']['y_average']['value']),
                                tol_pos=int(data['Arc Flash']['y_average']['pos_tolerance']),
                                tol_neg=int(data['Arc Flash']['y_average']['neg_tolerance']))
            
            
            canvas = FigureCanvasTkAgg(fig, master=y_average_frame)
            canvases["Y Average"] = canvas  # <--- SAVE IT
            canvas.draw()
            canvas.get_tk_widget().grid(row=1,column=column_index)
            column_index=column_index+1
        if chart_metrics.get("Y Maximum"):
            fig=Figure(figsize=(fig_w,fig_h),dpi=100)
            axes["Y Maximum"] = fig.add_subplot(111)
            ax=axes["Y Maximum"]
            y_maximum_frame = ctk.CTkFrame(y_position_frame)
            y_maximum_frame.grid(row=1,column=column_index, sticky = "w", padx=10, pady=5)
            # Set title and labels ONCE
            ax.set_title("Y Maximum")
            ax.set_xlabel("Frame Index")
            ax.set_ylabel("Pixel Value")

            
            # Now call plot_feature_on_axes without title/xlabel/ylabel every time
            plot_feature_on_axes(ax, "Solidification Zone",
                                x_data=[0, 1, 2], y_data=[0, 1, 2],
                                desired_value=int(data['Solidification Zone']['y_max']['value']),
                                tol_pos=int(data['Solidification Zone']['y_max']['pos_tolerance']),
                                tol_neg=int(data['Solidification Zone']['y_max']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Welding Wire",
                                x_data=[0, 1, 2], y_data=[0, 2, 4],
                                desired_value=int(data['Welding Wire']['y_max']['value']),
                                tol_pos=int(data['Welding Wire']['y_max']['pos_tolerance']),
                                tol_neg=int(data['Welding Wire']['y_max']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Arc Flash",
                                x_data=[0, 1, 2], y_data=[0, 3, 6],
                                desired_value=int(data['Arc Flash']['y_max']['value']),
                                tol_pos=int(data['Arc Flash']['y_max']['pos_tolerance']),
                                tol_neg=int(data['Arc Flash']['y_max']['neg_tolerance']))
            canvas = FigureCanvasTkAgg(fig, master=y_maximum_frame)
            canvases["Y Maximum"] = canvas  # <--- SAVE IT
            canvas.draw()
            canvas.get_tk_widget().grid(row=1,column=column_index)
            column_index=column_index+1
            
        if chart_metrics.get("Y Minimum"):
            fig=Figure(figsize=(fig_w,fig_h),dpi=100)
            axes["Y Minimum"] = fig.add_subplot(111)
            ax=axes["Y Minimum"]
            y_minimum_frame = ctk.CTkFrame(y_position_frame)
            y_minimum_frame.grid(row=1,column=column_index, sticky = "w", padx=10, pady=5)
            ax.set_title("Y Minimum")
            ax.set_xlabel("Frame Index")
            ax.set_ylabel("Pixel Value")

            
            # Now call plot_feature_on_axes without title/xlabel/ylabel every time
            plot_feature_on_axes(ax, "Solidification Zone",
                                x_data=[0, 1, 2], y_data=[0, 1, 2],
                                desired_value=int(data['Solidification Zone']['y_min']['value']),
                                tol_pos=int(data['Solidification Zone']['y_min']['pos_tolerance']),
                                tol_neg=int(data['Solidification Zone']['y_min']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Welding Wire",
                                x_data=[0, 1, 2], y_data=[0, 2, 4],
                                desired_value=int(data['Welding Wire']['y_min']['value']),
                                tol_pos=int(data['Welding Wire']['y_min']['pos_tolerance']),
                                tol_neg=int(data['Welding Wire']['y_min']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Arc Flash",
                                x_data=[0, 1, 2], y_data=[0, 3, 6],
                                desired_value=int(data['Arc Flash']['y_min']['value']),
                                tol_pos=int(data['Arc Flash']['y_min']['pos_tolerance']),
                                tol_neg=int(data['Arc Flash']['y_min']['neg_tolerance']))
            canvas = FigureCanvasTkAgg(fig, master=y_minimum_frame)
            canvases["Y Minimum"] = canvas  # <--- SAVE IT
            canvas.draw()
            canvas.get_tk_widget().grid(row=1,column=column_index)
            column_index=+1
    
    if show_charts.get("Position Standard Deviations"):
        column_index = 0
        position_std_frame = ctk.CTkFrame(chart_frame)
        position_std_frame_label = ctk.CTkLabel(position_std_frame, text= "Position Standard Deviation", font=("Arial", 12, "bold"))
        position_std_frame.grid(row=0, column=1, padx=20, pady=20)
        position_std_frame_label.grid(row=0,column=0,sticky="w", padx=10, pady=5)
        
        if chart_metrics.get("X Average Standard Deviation"):
            fig=Figure(figsize=(fig_w,fig_h),dpi=100)
            axes["X Average Standard Deviation"] = fig.add_subplot(111)
            ax=axes["X Average Standard Deviation"]
            x_avg_std_dev_frame = ctk.CTkFrame(position_std_frame)
            x_avg_std_dev_frame.grid(row=1,column=column_index, sticky = "w", padx=10, pady=5)
            ax.set_title("X Average Standard Deviation")
            ax.set_xlabel("Frame Index")
            ax.set_ylabel("Pixel Value")

            
            # Now call plot_feature_on_axes without title/xlabel/ylabel every time
            plot_feature_on_axes(ax, "Solidification Zone",
                                x_data=[0, 1, 2], y_data=[0, 1, 2],
                                desired_value=int(data['Solidification Zone']['x_avg_std_deviation']['value']),
                                tol_pos=int(data['Solidification Zone']['x_avg_std_deviation']['pos_tolerance']),
                                tol_neg=int(data['Solidification Zone']['x_avg_std_deviation']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Welding Wire",
                                x_data=[0, 1, 2], y_data=[0, 2, 4],
                                desired_value=int(data['Welding Wire']['x_avg_std_deviation']['value']),
                                tol_pos=int(data['Welding Wire']['x_avg_std_deviation']['pos_tolerance']),
                                tol_neg=int(data['Welding Wire']['x_avg_std_deviation']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Arc Flash",
                                x_data=[0, 1, 2], y_data=[0, 3, 6],
                                desired_value=int(data['Arc Flash']['x_avg_std_deviation']['value']),
                                tol_pos=int(data['Arc Flash']['x_avg_std_deviation']['pos_tolerance']),
                                tol_neg=int(data['Arc Flash']['x_avg_std_deviation']['neg_tolerance']))
            canvas = FigureCanvasTkAgg(fig, master=x_avg_std_dev_frame)
            canvas.draw()
            canvases["X Average Standard Deviation"] = canvas  # <--- SAVE IT
            canvas.get_tk_widget().grid(row=1,column=column_index)
            column_index=column_index+1
            
        if chart_metrics.get("Y Average Standard Deviation"):
            fig=Figure(figsize=(fig_w,fig_h),dpi=100)
            axes["Y Average Standard Deviation"] = fig.add_subplot(111)
            ax=axes["Y Average Standard Deviation"]
            y_avg_std_dev_frame = ctk.CTkFrame(position_std_frame)
            y_avg_std_dev_frame.grid(row=1,column=column_index, sticky = "w", padx=10, pady=5)
            ax.set_title("Y Average Standard Deviation")
            ax.set_xlabel("Frame Index")
            ax.set_ylabel("Pixel Value")

            
            # Now call plot_feature_on_axes without title/xlabel/ylabel every time
            plot_feature_on_axes(ax, "Solidification Zone",
                                x_data=[0, 1, 2], y_data=[0, 1, 2],
                                desired_value=int(data['Solidification Zone']['y_avg_std_deviation']['value']),
                                tol_pos=int(data['Solidification Zone']['y_avg_std_deviation']['pos_tolerance']),
                                tol_neg=int(data['Solidification Zone']['y_avg_std_deviation']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Welding Wire",
                                x_data=[0, 1, 2], y_data=[0, 2, 4],
                                desired_value=int(data['Welding Wire']['y_avg_std_deviation']['value']),
                                tol_pos=int(data['Welding Wire']['y_avg_std_deviation']['pos_tolerance']),
                                tol_neg=int(data['Welding Wire']['y_avg_std_deviation']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Arc Flash",
                                x_data=[0, 1, 2], y_data=[0, 3, 6],
                                desired_value=int(data['Arc Flash']['y_avg_std_deviation']['value']),
                                tol_pos=int(data['Arc Flash']['y_avg_std_deviation']['pos_tolerance']),
                                tol_neg=int(data['Arc Flash']['y_avg_std_deviation']['neg_tolerance']))
            canvas = FigureCanvasTkAgg(fig, master=x_avg_std_dev_frame)
            canvases["Y Average Standard Deviation"] = canvas  # <--- SAVE IT
            canvas.draw()
            canvas.get_tk_widget().grid(row=1,column=column_index)
            column_index=column_index+1
            
      
        
    if show_charts.get("Class Area"):
        column_index = 0
        class_area_frame = ctk.CTkFrame(chart_frame)
        class_area_frame_label = ctk.CTkLabel(class_area_frame, text= "Class Area and Standard Deviation", font=("Arial", 12, "bold"))
        class_area_frame.grid(row=1, column=1, padx=20, pady=20)
        class_area_frame_label.grid(row=0,column=0,sticky="w", padx=10, pady=5)
        if chart_metrics.get("Class Area"):
            fig=Figure(figsize=(fig_w,fig_h),dpi=100)
            axes["Class Area"] = fig.add_subplot(111)
            ax=axes["Class Area"]
            class_area_sub_frame = ctk.CTkFrame(class_area_frame)
            class_area_sub_frame.grid(row=1,column=column_index, sticky = "w", padx=10, pady=5)
            ax.set_title("Class Area")
            ax.set_xlabel("Frame Index")
            ax.set_ylabel("Pixel Value")

            
            # Now call plot_feature_on_axes without title/xlabel/ylabel every time
            plot_feature_on_axes(ax, "Solidification Zone",
                                x_data=[0, 1, 2], y_data=[0, 1, 2],
                                desired_value=int(data['Solidification Zone']['area_average']['value']),
                                tol_pos=int(data['Solidification Zone']['area_average']['pos_tolerance']),
                                tol_neg=int(data['Solidification Zone']['area_average']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Welding Wire",
                                x_data=[0, 1, 2], y_data=[0, 2, 4],
                                desired_value=int(data['Welding Wire']['area_average']['value']),
                                tol_pos=int(data['Welding Wire']['area_average']['pos_tolerance']),
                                tol_neg=int(data['Welding Wire']['area_average']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Arc Flash",
                                x_data=[0, 1, 2], y_data=[0, 3, 6],
                                desired_value=int(data['Arc Flash']['area_average']['value']),
                                tol_pos=int(data['Arc Flash']['area_average']['pos_tolerance']),
                                tol_neg=int(data['Arc Flash']['area_average']['neg_tolerance']))
            canvas = FigureCanvasTkAgg(fig, master=class_area_sub_frame)
            canvas.draw()
            canvases["Class Area"] = canvas  # <--- SAVE IT
            canvas.get_tk_widget().grid(row=1,column=column_index)
            column_index=column_index+1
            
    if show_charts.get("Class Area Standard Deviation"):
        if chart_metrics.get("Class Area"):
            fig=Figure(figsize=(fig_w,fig_h),dpi=100)
            axes["Class Area Standard Deviation"] = fig.add_subplot(111)
            ax=axes["Class Area Standard Deviation"]
            class_std_dev_sub_frame = ctk.CTkFrame(class_area_frame)
            class_std_dev_sub_frame.grid(row=1,column=column_index, sticky = "w", padx=10, pady=5)
            ax.set_title("Class Area Standard Deviation")
            ax.set_xlabel("Frame Index")
            ax.set_ylabel("Pixel Value")

            
            # Now call plot_feature_on_axes without title/xlabel/ylabel every time
            plot_feature_on_axes(ax, "Solidification Zone",
                                x_data=[0, 1, 2], y_data=[0, 1, 2],
                                desired_value=int(data['Solidification Zone']['area_std_deviation']['value']),
                                tol_pos=int(data['Solidification Zone']['area_std_deviation']['pos_tolerance']),
                                tol_neg=int(data['Solidification Zone']['area_std_deviation']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Welding Wire",
                                x_data=[0, 1, 2], y_data=[0, 2, 4],
                                desired_value=int(data['Welding Wire']['area_std_deviation']['value']),
                                tol_pos=int(data['Welding Wire']['area_std_deviation']['pos_tolerance']),
                                tol_neg=int(data['Welding Wire']['area_std_deviation']['neg_tolerance']))
            
            plot_feature_on_axes(ax, "Arc Flash",
                                x_data=[0, 1, 2], y_data=[0, 3, 6],
                                desired_value=int(data['Arc Flash']['area_std_deviation']['value']),
                                tol_pos=int(data['Arc Flash']['area_std_deviation']['pos_tolerance']),
                                tol_neg=int(data['Arc Flash']['area_std_deviation']['neg_tolerance']))
            canvas = FigureCanvasTkAgg(fig, master=class_std_dev_sub_frame)
            canvases["Class Area Standard Deviation"] = canvas  # <--- SAVE IT
            canvas.draw()
            canvas.get_tk_widget().grid(row=1,column=column_index)
            column_index=column_index+1
        

   
    


# Graphical viewing

chart_window = ctk.CTkToplevel()
chart_window.title("Charts Display")
chart_window.geometry("800x600")
# Chart container
chart_frame = ctk.CTkFrame(chart_window)
chart_frame.pack(fill="both", expand=True, padx=20, pady=20)
    
    
render_charts()
    


def detect_basler_camera():
    try:
        tl_factory = pylon.TlFactory.GetInstance()
        devices = tl_factory.EnumerateDevices()
        return devices[0] if devices else None
    except Exception as e:
        print(f"Error detecting camera: {e}")
        return None

def update_status_loop(label):
    global cam, camera
    connected_once = False

    while not connected_once:
        # Show "Attempting to connect"
        label.after(0, lambda: label.configure(text="Attempting to connect", text_color="yellow"))

        device = detect_basler_camera()

        if device:
            try:
                cam_obj = pylon.InstantCamera(pylon.TlFactory.GetInstance().CreateDevice(device))
                cam_obj.Open()
                camera = cam_obj  # Store for later use
                # camera.AcquisitionFrameRateEnable.SetValue(True)
                
                # camera.AcquisitionFrameRate.SetValue(20.0)
                connected_once = True

                # Update label to Connected
                label.after(0, lambda: label.configure(text="Connected", text_color="green"))

                with cam_lock:
                    cam = 1

                print(f"✅ Camera connected: {camera.GetDeviceInfo().GetModelName()}")
            except Exception as e:
                print(f"❌ Failed to open camera: {e}")
                label.after(0, lambda: label.configure(text="Disconnected", text_color="red"))
                with cam_lock:
                    cam = 0
        else:
            label.after(0, lambda: label.configure(text="Disconnected", text_color="red"))
            with cam_lock:
                cam = 0

        if not connected_once:
            time.sleep(2)
    

thread = threading.Thread(target=update_status_loop, args=(big_camera_display,), daemon=True)
thread.start()


#run
window.mainloop()
